import json
from openpyxl import Workbook

# Укажите путь к файлу JSON с двойными обратными слешами или префиксом `r`
json_file_path = r"C:\Users\Downloads\Untitled-5.json"

# Чтение JSON из файла
with open(json_file_path, "r") as json_file:
    json_data = json.load(json_file)

# Получаем значения hosts.host и hosts.interfaces.ip
hosts = json_data["zabbix_export"]["hosts"]

# Создание нового Excel-файла
workbook = Workbook()
sheet = workbook.active

# Заголовки столбцов
sheet["A1"] = "Host Name"
sheet["B1"] = "IP Address"

# Запись данных в файл
row_index = 2
for host in hosts:
    host_name = host["host"]
    interfaces = host["interfaces"]
    
    for interface in interfaces:
        ip = interface["ip"]
        
        sheet.cell(row=row_index, column=1, value=host_name)
        sheet.cell(row=row_index, column=2, value=ip)
        
        row_index += 1

# Сохранение файла
xls_file_path = r"C:\Users\Downloads\zbx_export_hosts.xlsx"
workbook.save(xls_file_path)

print(f"Результаты экспортированы в файл: {xls_file_path}")
